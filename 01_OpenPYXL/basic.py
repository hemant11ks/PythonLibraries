import openpyxl
from openpyxl import Workbook

wb = openpyxl.load_workbook('data.xlsx')
ws = wb.active
# It will print the name of worksheet
# print(ws)

# Setting value in cell A1
# ws['A1'] = 'Hi Buddy'

# Printing the value in cell A1
# print(ws['A1'].value)

# ws['A1'] = 'value': Set a value in cell A1.
# ws.append(row): Append a row of values to the worksheet.
# ws.insert_rows(1): Insert a new row at the beginning of the worksheet.
# ws.delete_rows(1): Delete a row from the worksheet.
# ws.merge_cells('A1:B1'): Merge cells A1 and B1.
# ws.unmerge_cells('A1:B1'): Unmerge cells A1 and B1.

# Writing data to call


wb = openpyxl.load_workbook('data.xlsx')
ws = wb.active

data = [
    ['Name', 'Age'],
    ['John', 21],
    ['Kumar', 16]
]

for row in data:
    ws.append(row)

wb.save('data.xlsx')
