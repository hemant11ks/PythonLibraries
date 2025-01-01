from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# wb = load_workbook('File.xlsx')
# ws = wb.active
# Or I can use: ws = wb['Sheet']
# ws['A1'].value = "test"
# print(ws['A1'].value)

# print(wb.sheetnames)
#
# print(wb['Sheet'])
#
# wb.save('File.xlsx')

# Now creating the sheet ourself

# wb = load_workbook('File.xlsx')
# wb.create_sheet("Test")
# print(wb.sheetnames)

# Creating a new workbook and performing operations

# wb = Workbook() # It will initialize the new workbook
# ws = wb.active # Default sheet
# ws.title = "Data"

# This is not the efficient way for writing the data
# ws["A1"] = "dummy"

# So we can use the Adding/Appending rows
# ws.append(['Hemant', 'is', 'my', 'name'])
# ws.append(['Kumar', 'is', 'my', 'name'])
# ws.append(['Singh', 'is', 'my', 'name'])

# Accessing multiple cells / Looping

wb = load_workbook('File.xlsx')
ws = wb.active

# For col we required A1, B1 that why we using the openpyxl inbuilt functions : combining the character with the row
# for row in range(1,4):
#     for col in range(1,4):
#         char = get_column_letter(col)
#         print(ws[char + str(row)].value)
# If we want to mark the value of cell according to the name we can use this like:
# ws[char + str(row)] = char + str(row)

# For merging and de merging the data: Keep not that you will lose the data
# ws.merged_cells("A1:D1")
# ws.unmerged_cells("A1:D1")

# Insert and Delete rows/columns

# ws.insert_cols(7) # We can give int here
# ws.insert_rows(7)
# ws.delete_cols(7)
# ws.delete_rows(7)




# wb.save('Hemant.xlsx')



